from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"D:\Gymlog dokumentointi\Gainer Program\gainer-programs.xlsx")
TARGET = ROOT / "src" / "features" / "workout" / "gainerProgramCatalog.ts"

PROGRESSION_RULES = {
    "primary": "Use double progression on anchor lifts. Add load after the top of the rep range is repeatable with clean form.",
    "secondary": "Progress secondary lifts by adding reps first, then load once the range is stable across working sets.",
    "accessory": "Keep accessories controlled. Add reps before load and avoid form breakdown near fatigue.",
    "failureHandling": "If repsMin is missed, repeat the load next time. If it happens twice, reduce load 5-10% and rebuild.",
}

PROGRAM_NAME_OVERRIDES = {
    "beginner_bro_split": "Bro Split",
}


def clean_text(value: object) -> str:
    text = str(value or "").strip()
    text = text.replace("—", "-").replace("–", "-")
    text = text.replace("5�5", "5x5").replace("5×5", "5x5")
    text = text.replace("At Home � No Equipment", "At Home - No Equipment")
    text = text.replace("�", "-")
    return text


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug or "item"


def ts(value: object) -> str:
    return json.dumps(value, ensure_ascii=False)


def normalize_level(value: str) -> str:
    lowered = value.lower()
    if "advanced" in lowered or "expert" in lowered:
        return "advanced"
    if "intermediate" in lowered:
        return "intermediate"
    return "beginner"


def normalize_goal(value: str) -> str:
    lowered = value.lower()
    if "strong" in lowered or "strength" in lowered:
        return "strength"
    if any(token in lowered for token in ["muscle", "sculpt", "glute"]):
        return "hypertrophy"
    return "general"


def normalize_split(days_per_week: int) -> str:
    if days_per_week <= 3:
        return "full_body"
    if days_per_week == 4:
        return "upper_lower"
    return "hybrid"


def parse_muscles(value: object) -> list[str]:
    return [item.strip().lower() for item in clean_text(value).split(",") if item and item.strip()]


def resolve_group(name: str, muscles: list[str], equipment: str) -> str:
    lowered_name = name.lower()
    joined = " ".join(muscles)

    if name == "Dynamic Warm-Up":
        return "gainer_warmup"
    if name == "Cooldown Flow":
        return "gainer_cooldown"
    if any(token in joined or token in lowered_name for token in ["core", "abs", "plank", "crunch"]):
        return "gainer_core"
    if "calf" in joined or "calf" in lowered_name or "calves" in joined:
        return "gainer_calves"
    if any(token in joined or token in lowered_name for token in ["biceps", "triceps", "forearm", "curl", "pushdown", "skull"]):
        return "gainer_arms"
    if any(token in joined or token in lowered_name for token in ["delt", "shoulder", "lateral raise", "arnold press", "face pull"]):
        return "gainer_shoulders"
    if any(token in joined or token in lowered_name for token in ["chest", "bench", "push-up", "dip", "crossover", "fly"]):
        return "gainer_press_chest"
    if any(token in joined or token in lowered_name for token in ["lat", "back", "row", "pull-up", "pulldown", "pullover"]):
        return "gainer_pull_back"
    if any(token in joined or token in lowered_name for token in ["glute", "hamstring", "deadlift", "thrust", "bridge", "kickback", "leg curl"]):
        return "gainer_hinge_glute"
    if any(token in joined or token in lowered_name for token in ["quad", "leg", "squat", "lunge", "step-up", "adductor"]):
        return "gainer_squat_leg"
    if any(token in joined or token in lowered_name for token in ["run", "hiit", "mobility", "stretch", "yoga", "burpee", "conditioning"]):
        return "gainer_conditioning"
    if equipment == "bodyweight":
        return "gainer_bodyweight"
    return "gainer_general"


def resolve_tracking_mode(equipment: str, rest_seconds: int, role: str) -> str:
    if equipment == "bodyweight" or rest_seconds == 0:
        return "bodyweight"
    if role == "accessory":
        return "reps_first"
    return "load_and_reps"


def resolve_role(order: int) -> tuple[str, str]:
    if order == 1:
        return "primary", "high"
    if order <= 3:
        return "secondary", "medium"
    return "accessory", "low"


def rest_bounds(rest_seconds: int) -> tuple[int, int]:
    if rest_seconds <= 0:
        return 0, 0
    return max(30, rest_seconds - 30), rest_seconds + 30


def read_workbook(path: Path):
    workbook = load_workbook(path, data_only=True)
    programs_sheet = workbook["Programs"]
    workouts_sheet = workbook["Workouts"]
    exercises_sheet = workbook["Exercises"]

    programs = {}
    for row in programs_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        program_id = clean_text(row[0])
        title = PROGRAM_NAME_OVERRIDES.get(program_id, clean_text(row[1]))
        programs[program_id] = {
            "sourceId": program_id,
            "id": f"tpl_gainer_{program_id}_v1",
            "name": title,
            "level": normalize_level(clean_text(row[2])),
            "goalType": normalize_goal(clean_text(row[3])),
            "daysPerWeek": int(row[5] or 1),
            "estimatedSessionDuration": 0,
            "sessions": [],
        }

    workout_rows = {}
    for row in workouts_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        program_id = clean_text(row[0])
        day = int(row[1])
        workout_rows[(program_id, day)] = {
            "id": slugify(clean_text(row[2])),
            "name": clean_text(row[2]),
            "orderIndex": day,
            "estimatedDuration": int(row[3] or 0),
            "exercises": [],
        }

    for row in exercises_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        program_id = clean_text(row[0])
        day = int(row[1])
        order = int(row[2])
        name = clean_text(row[3])
        muscles = parse_muscles(row[4])
        equipment = slugify(clean_text(row[5]))
        sets = int(row[6] or 1)
        reps_min = int(row[7] or 1)
        reps_max = int(row[8] or reps_min)
        rest_seconds = int(row[9] or 0)
        role, priority = resolve_role(order)
        group = resolve_group(name, muscles, equipment)
        rest_min, rest_max = rest_bounds(rest_seconds)

        workout = workout_rows.get((program_id, day))
        if workout is None:
            continue

        workout["exercises"].append(
            {
                "id": slugify(f"{workout['id']}_{name}"),
                "exerciseName": name,
                "slotId": f"{role}_{group}_{order}",
                "role": role,
                "progressionPriority": priority,
                "trackingMode": resolve_tracking_mode(equipment, rest_seconds, role),
                "sets": sets,
                "repsMin": reps_min,
                "repsMax": reps_max,
                "restSecondsMin": rest_min,
                "restSecondsMax": rest_max,
                "substitutionGroup": group,
            }
        )

    for (program_id, _day), workout in sorted(workout_rows.items(), key=lambda item: (item[0][0], item[0][1])):
        program = programs.get(program_id)
        if not program:
            continue

        session_slug = workout["id"]
        warmup = {
            "id": f"{session_slug}_dynamic_warm_up",
            "exerciseName": "Dynamic Warm-Up",
            "slotId": "warmup_gainer_warmup_0",
            "role": "accessory",
            "progressionPriority": "low",
            "trackingMode": "bodyweight",
            "sets": 1,
            "repsMin": 5,
            "repsMax": 8,
            "restSecondsMin": 30,
            "restSecondsMax": 45,
            "substitutionGroup": "gainer_warmup",
        }
        cooldown = {
            "id": f"{session_slug}_cooldown_flow",
            "exerciseName": "Cooldown Flow",
            "slotId": "finish_gainer_cooldown_99",
            "role": "accessory",
            "progressionPriority": "low",
            "trackingMode": "bodyweight",
            "sets": 1,
            "repsMin": 3,
            "repsMax": 5,
            "restSecondsMin": 0,
            "restSecondsMax": 0,
            "substitutionGroup": "gainer_cooldown",
        }
        workout["exercises"] = [warmup, *workout["exercises"], cooldown]
        program["sessions"].append(
            {
                "id": workout["id"],
                "name": workout["name"],
                "orderIndex": workout["orderIndex"],
                "exercises": workout["exercises"],
            }
        )
        program["estimatedSessionDuration"] = max(program["estimatedSessionDuration"], workout["estimatedDuration"])

    return list(programs.values())


def build_groups(programs: list[dict]) -> list[dict]:
    names_by_group: dict[str, set[str]] = {
        "gainer_warmup": {"Dynamic Warm-Up", "Mobility Warm-Up", "Joint Prep Flow"},
        "gainer_cooldown": {"Cooldown Flow", "Recovery Stretch", "Breathing Reset"},
    }

    for program in programs:
        for session in program["sessions"]:
            for exercise in session["exercises"]:
                names_by_group.setdefault(exercise["substitutionGroup"], set()).add(exercise["exerciseName"])

    return [
        {"id": group_id, "allowedExerciseNames": sorted(names)}
        for group_id, names in sorted(names_by_group.items())
    ]


def render_exercise(exercise: dict, indent: str) -> str:
    ordered_keys = [
        "id",
        "exerciseName",
        "slotId",
        "role",
        "progressionPriority",
        "trackingMode",
        "sets",
        "repsMin",
        "repsMax",
        "restSecondsMin",
        "restSecondsMax",
        "substitutionGroup",
    ]
    body = ", ".join(f"{key}: {ts(exercise[key])}" for key in ordered_keys)
    return f"{indent}{{ {body} }}"


def render_catalog(programs: list[dict], groups: list[dict], tail: str) -> str:
    lines = [
        "import type { WorkoutSubstitutionGroup, WorkoutTemplateV1 } from './workoutTypes';",
        "import type { RecommendationProgramDefinition } from '../../types/recommendation';",
        "",
        "const GAINER_PROGRESSION_RULES = {",
    ]
    for key, value in PROGRESSION_RULES.items():
        lines.append(f"  {key}: {ts(value)},")
    lines.extend([
        "} as const;",
        "",
        "export const GAINER_PROGRAM_SUBSTITUTION_GROUPS: WorkoutSubstitutionGroup[] = [",
    ])
    for group in groups:
        lines.append(f"  {{ id: {ts(group['id'])}, allowedExerciseNames: {ts(group['allowedExerciseNames'])} }},")
    lines.extend([
        "];",
        "",
        "export const GAINER_WORKOUT_TEMPLATES_V1: WorkoutTemplateV1[] = [",
    ])

    for program in programs:
        lines.extend([
            "  {",
            f"    id: {ts(program['id'])},",
            f"    name: {ts(program['name'])},",
            f"    goalType: {ts(program['goalType'])},",
            f"    level: {ts(program['level'])},",
            f"    splitType: {ts(normalize_split(program['daysPerWeek']))},",
            f"    daysPerWeek: {program['daysPerWeek']},",
            f"    estimatedSessionDuration: {program['estimatedSessionDuration']},",
            "    progressionModel: 'double_progression',",
            "    defaultScheduleMode: 'rolling_sequence',",
            "    progressionRules: GAINER_PROGRESSION_RULES,",
            "    sessions: [",
        ])
        for session in program["sessions"]:
            lines.extend([
                f"      {{ id: {ts(session['id'])}, name: {ts(session['name'])}, orderIndex: {session['orderIndex']}, exercises: [",
            ])
            for exercise in session["exercises"]:
                lines.append(render_exercise(exercise, "        ") + ",")
            lines.append("      ] },")
        lines.extend([
            "    ],",
            "  },",
        ])
    lines.extend([
        "];",
        "",
        tail.strip(),
        "",
    ])
    return "\n".join(lines)


def main() -> None:
    source = DEFAULT_XLSX
    current = TARGET.read_text(encoding="utf-8")
    tail_marker = "type GainerRecommendationConfig"
    tail_start = current.index(tail_marker)
    tail = current[tail_start:]
    programs = read_workbook(source)
    groups = build_groups(programs)
    TARGET.write_text(render_catalog(programs, groups, tail), encoding="utf-8")
    print(f"Generated {len(programs)} Gainer programs from {source}")


if __name__ == "__main__":
    main()
